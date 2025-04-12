import json
import requests
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.http import HttpResponse, JsonResponse
from django.contrib.auth import get_user_model
CustomUser = get_user_model()
from django.shortcuts import get_object_or_404, redirect, render, reverse
from django.views.decorators.csrf import csrf_exempt
from .models import Attendance, AttendanceReport, Student, Subject, Session
from .EmailBackend import EmailBackend
import logging
from datetime import datetime
from django.shortcuts import get_object_or_404, redirect
from django.http import JsonResponse
from django.contrib import messages
from .models import Student, AttendanceReport, LeaveReportStudent, FeedbackStudent, NotificationStudent 
from django.contrib.auth.decorators import login_required
from .models import FeedbackStudent as Feedback
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.http import HttpResponse
from django.urls import path
from datetime import datetime

# Create your views here.

def login_page(request):
    if request.user.is_authenticated:
        if request.user.user_type == '1':
            return redirect(reverse("admin_home"))
        elif request.user.user_type == '2':
            return redirect(reverse("staff_home"))
        else:
            return redirect(reverse("student_home"))
    return render(request, 'main_app/login.html')


def doLogin(request, **kwargs):
    if request.method != 'POST':
        return HttpResponse("<h4>Denied</h4>")
    else:
        # Authenticate
        user = EmailBackend.authenticate(request, username=request.POST.get('email'), password=request.POST.get('password'))
        if user is not None:
            login(request, user)
            if user.user_type == '1':
                return redirect(reverse("admin_home"))
            elif user.user_type == '2':
                return redirect(reverse("staff_home"))
            else:
                return redirect(reverse("student_home"))
        else:
            messages.error(request, "Invalid details")
            return redirect("/")

def logout_user(request):
    if request.user is not None:
        logout(request)
    return redirect("/")

def export_students_excel(request):
    # Create a new workbook and select the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students Report"
    
    # Define styles
    header_font = Font(name='Arial', bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    # Add headers
    headers = ['#', 'Full Name', 'Email', 'Gender', 'Course', 'Student ID']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Add data
    students = CustomUser.objects.filter(user_type=3)  # Assuming 3 is the user_type for students
    for row_num, student in enumerate(students, 2):
        # Row number
        ws.cell(row=row_num, column=1).value = row_num - 1
        
        # Full Name
        ws.cell(row=row_num, column=2).value = f"{student.first_name}, {student.last_name}"
        
        # Email
        ws.cell(row=row_num, column=3).value = student.email
        
        # Gender
        ws.cell(row=row_num, column=4).value = student.gender
        
        # Course
        ws.cell(row=row_num, column=5).value = student.student.course.name
        
        # Student ID
        ws.cell(row=row_num, column=6).value = student.student_code if hasattr(student, 'student_code') else ""
        
        # Apply borders to all cells in this row
        for col_num in range(1, 7):
            ws.cell(row=row_num, column=col_num).border = thin_border
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    # Create the HttpResponse with Excel content type
    filename = f"Students_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Save the workbook to the response
    wb.save(response)
    return response
def export_attendance_excel(request):
    # Get parameters from request
    subject_id = request.GET.get('subject')
    session_id = request.GET.get('session')
    attendance_date_id = request.GET.get('attendance_date_id')
    
    # Validate parameters
    if not all([subject_id, session_id, attendance_date_id]):
        return HttpResponse("Missing parameters", status=400)
    
    try:
        # Get necessary data
        subject = Subject.objects.get(id=subject_id)
        session = Session.objects.get(id=session_id)
        attendance = Attendance.objects.get(id=attendance_date_id)
        attendance_reports = AttendanceReport.objects.filter(attendance=attendance)
        
        # Create a new workbook and select the active worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance Report"
        
        # Define styles
        header_font = Font(name='Arial', bold=True, size=12, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        present_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        absent_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        # Add title and info
        ws.merge_cells('A1:D1')
        ws.cell(row=1, column=1).value = "Attendance Report"
        ws.cell(row=1, column=1).font = Font(name='Arial', bold=True, size=14)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A2:D2')
        ws.cell(row=2, column=1).value = f"Subject: {subject.name} | Session: {session} | Date: {attendance.attendance_date}"
        ws.cell(row=2, column=1).font = Font(name='Arial', italic=True)
        ws.cell(row=2, column=1).alignment = Alignment(horizontal='center')
        
        # Add headers (row 4)
        headers = ['#', 'Student Name', 'Status']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Add data
        present_count = 0
        absent_count = 0
        
        for row_num, report in enumerate(attendance_reports, 5):
            # Row number
            cell = ws.cell(row=row_num, column=1)
            cell.value = row_num - 4
            cell.border = thin_border
            
            # Student Name
            cell = ws.cell(row=row_num, column=2)
            cell.value = f"{report.student.admin.first_name} {report.student.admin.last_name}"
            cell.border = thin_border
            
            # Status
            cell = ws.cell(row=row_num, column=3)
            status = "Present" if report.status else "Absent"
            cell.value = status
            cell.border = thin_border
            
            # Color coding for status
            if report.status:
                cell.fill = present_fill
                present_count += 1
            else:
                cell.fill = absent_fill
                absent_count += 1
        
        # Add summary
        summary_row = len(attendance_reports) + 5
        ws.merge_cells(f'A{summary_row}:B{summary_row}')
        ws.cell(row=summary_row, column=1).value = f"Total Students: {len(attendance_reports)}"
        ws.cell(row=summary_row, column=1).font = Font(bold=True)
        ws.cell(row=summary_row, column=1).border = thin_border
        
        ws.cell(row=summary_row, column=3).value = f"Present: {present_count} | Absent: {absent_count}"
        ws.cell(row=summary_row, column=3).font = Font(bold=True)
        ws.cell(row=summary_row, column=3).border = thin_border
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        
        # Create the HttpResponse with Excel content type
        filename = f"Attendance_Report_{subject.name}_{attendance.attendance_date.strftime('%Y%m%d')}.xlsx"
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Save the workbook to the response
        wb.save(response)
        return response
        
    except Exception as e:
        return HttpResponse(f"Error: {str(e)}", status=500)

def delete_student(request, student_id):
    try:
        student = get_object_or_404(Student, id=student_id)
        user = student.admin

        # Delete related AttendanceReport records
        AttendanceReport.objects.filter(student=student).delete()

        # Delete related LeaveReportStudent records
        LeaveReportStudent.objects.filter(student=student).delete()

        # Delete related FeedbackStudent records
        FeedbackStudent.objects.filter(student=student).delete()

        # Delete related NotificationStudent records
        NotificationStudent.objects.filter(student=student).delete()

        # Delete the student and user records
        student.delete()
        user.delete()

        messages.success(request, "Student deleted successfully!")
        return JsonResponse({'message': 'Student deleted successfully'})
    except Exception as e:
        messages.error(request, f"Error deleting student: {str(e)}")
        return JsonResponse({'message': f"Error deleting student: {str(e)}"}, status=500)
    return redirect('manage_student')

@csrf_exempt
def get_attendance(request):
    subject_id = request.POST.get('subject')
    session_id = request.POST.get('session')
    try:
        subject = get_object_or_404(Subject, id=subject_id)
        session = get_object_or_404(Session, id=session_id)
        attendance = Attendance.objects.filter(subject=subject, session=session)
        attendance_list = []
        for attd in attendance:
            data = {
                "id": attd.id,
                "attendance_date": str(attd.date),
                "session": attd.session.id
            }
            attendance_list.append(data)
        return JsonResponse(json.dumps(attendance_list), safe=False)
    except Exception as e:
        return None


@csrf_exempt
def save_attendance_data(request):
    if request.method == 'POST':
        subject_id = request.POST.get('subject_id')
        session_id = request.POST.get('session_id')
        attendance_date = request.POST.get('attendance_date')
        student_ids = request.POST.getlist('student_ids[]')

        logging.info(f"Subject ID: {subject_id}")
        logging.info(f"Session ID: {session_id}")
        logging.info(f"Attendance Date: {attendance_date}")
        logging.info(f"Student IDs: {student_ids}")

        try:
            subject = get_object_or_404(Subject, id=subject_id)
            session = get_object_or_404(Session, id=session_id)

            # Convert attendance_date to a date object
            attendance_date_obj = datetime.strptime(attendance_date, '%Y-%m-%d').date()

            # Validate that the attendance date is within the session dates
            if not (session.start_year <= attendance_date_obj <= session.end_year):
                return JsonResponse({"status": "error", "message": "Attendance date is out of session range"})

            attendance = Attendance.objects.create(subject=subject, session=session, date=attendance_date_obj)

            for student_id in student_ids:
                logging.info(f"Processing student ID: {student_id}")
                try:
                    student = get_object_or_404(Student, id=student_id)
                    status = request.POST.get(f'status_{student_id}', 'False') == 'True'
                    AttendanceReport.objects.create(student=student, attendance=attendance, status=status)
                except Exception as e:
                    logging.error(f"Error processing student ID {student_id}: {str(e)}")
                    messages.error(request, f"Error processing student ID {student_id}: {str(e)}")
                    return JsonResponse({"status": "error", "message": f"Error processing student ID {student_id}: {str(e)}"})

            messages.success(request, "Attendance saved successfully!")
            return JsonResponse({"status": "success"})
        except Exception as e:
            logging.error(f"Error in saving attendance: {str(e)}")
            messages.error(request, f"Error in saving attendance: {str(e)}")
            return JsonResponse({"status": "error", "message": str(e)})

    return JsonResponse({"status": "error", "message": "Invalid request method"})



def manage_students(request):
    students = Student.objects.all()
    context = {
        'students': students,
    }
    return render(request, 'manage_students.html', context)

def showFirebaseJS(request):
    data = """
    // Give the service worker access to Firebase Messaging.
    // Note that you can only use Firebase Messaging here, other Firebase libraries
    // are not available in the service worker.
    importScripts('https://www.gstatic.com/firebasejs/7.22.1/firebase-app.js');
    importScripts('https://www.gstatic.com/firebasejs/7.22.1/firebase-messaging.js');

    // Initialize the Firebase app in the service worker by passing in
    // your app's Firebase config object.
    // https://firebase.google.com/docs/web/setup#config-object
    firebase.initializeApp({
        apiKey: "AIzaSyBarDWWHTfTMSrtc5Lj3Cdw5dEvjAkFwtM",
        authDomain: "sms-with-django.firebaseapp.com",
        databaseURL: "https://sms-with-django.firebaseio.com",
        projectId: "sms-with-django",
        storageBucket: "sms-with-django.appspot.com",
        messagingSenderId: "945324593139",
        appId: "1:945324593139:web:03fa99a8854bbd38420c86",
        measurementId: "G-2F2RXTL9GT"
    });

    // Retrieve an instance of Firebase Messaging so that it can handle background
    // messages.
    const messaging = firebase.messaging();
    messaging.setBackgroundMessageHandler(function (payload) {
        const notification = JSON.parse(payload);
        const notificationOption = {
            body: notification.body,
            icon: notification.icon
        }
        return self.registration.showNotification(payload.notification.title, notificationOption);
    });
    """
    return HttpResponse(data, content_type='application/javascript')

@login_required(login_url='user_login')
def student_feedback(request):
    feedbacks = Feedback.objects.filter(student_id=request.user.id)
    return render(request, 'main_app/student_feedback.html', {'feedbacks': feedbacks})

   