import json
from datetime import datetime

from django.contrib import messages
from django.core.files.storage import FileSystemStorage
from django.http import HttpResponse, JsonResponse
from django.shortcuts import (HttpResponseRedirect, get_object_or_404,redirect, render)
from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt
from django.http import Http404
from .forms import *
from .models import *
from django.views.decorators.http import require_http_methods
from django.utils import timezone


def staff_home(request):
    staff = get_object_or_404(Staff, admin=request.user)
    total_students = Student.objects.filter(course=staff.course).count()
    total_leave = LeaveReportStaff.objects.filter(staff=staff).count()
    subjects = Subject.objects.filter(staff=staff)
    total_subject = subjects.count()
    attendance_list = Attendance.objects.filter(subject__in=subjects)
    total_attendance = attendance_list.count()
    attendance_list = []
    subject_list = []
    for subject in subjects:
        attendance_count = Attendance.objects.filter(subject=subject).count()
        subject_list.append(subject.name)
        attendance_list.append(attendance_count)
    context = {
        'page_title': 'Staff Panel - ' + str(staff.admin.last_name) + ' (' + str(staff.course) + ')',
        'total_students': total_students,
        'total_attendance': total_attendance,
        'total_leave': total_leave,
        'total_subject': total_subject,
        'subject_list': subject_list,
        'attendance_list': attendance_list
    }
    return render(request, 'staff_template/home_content.html', context)

@csrf_exempt
def staff_take_attendance_by_qr(request):
    if request.method == 'POST':
        student_id = request.POST.get('student_id')
        subject_id = request.POST.get('subject_id')
        attendance_date = request.POST.get('attendance_date')
        # Get attendance status if provided, default to present (True) if not provided
        attendance_status = request.POST.get('attendance_status', 'True') == 'True'
        
        if not all([student_id, subject_id, attendance_date]):
            missing = []
            if not student_id: missing.append('student_id')
            if not subject_id: missing.append('subject_id')
            if not attendance_date: missing.append('attendance_date')
            return JsonResponse({
                'status': 'error',
                'message': f'Missing required fields: {", ".join(missing)}'
            })
        
        try:
            # First try to find student by student_code
            student = None
            try:
                student = Student.objects.select_related(
                    'admin', 
                    'session',
                    'course'
                ).get(admin__student_code=student_id)
            except Student.DoesNotExist:
                try:
                    student = Student.objects.select_related(
                        'admin', 
                        'session',
                        'course'
                    ).get(id=student_id)
                except Student.DoesNotExist:
                    return JsonResponse({
                        'status': 'error',
                        'message': f'No student found with ID or code: {student_id}'
                    })
            
            if not student.session:
                return JsonResponse({
                    'status': 'error',
                    'message': 'Student does not have a session assigned'
                })

            # Get the subject
            try:
                staff = Staff.objects.get(admin=request.user)
                subject = Subject.objects.select_related('course').get(
                    id=subject_id,
                    staff=staff  # This ensures the staff member can only take attendance for their subjects
                )
                
                print(f"[DEBUG] Student: {student.admin.first_name} {student.admin.last_name}")
                print(f"[DEBUG] Student Course: {student.course.name if student.course else 'None'}")
                print(f"[DEBUG] Subject: {subject.name}")
                print(f"[DEBUG] Subject Course: {subject.course.name if subject.course else 'None'}")
                
                # Check if attendance already exists for this student on this date
                attendance_date_obj = datetime.strptime(attendance_date, '%Y-%m-%d').date()
                
                # Verify that the attendance date falls within the student's session period
                if not (student.session.start_year <= attendance_date_obj <= student.session.end_year):
                    return JsonResponse({
                        'status': 'error',
                        'message': 'Attendance date is outside the student session period',
                        'details': f'Session period: {student.session.start_year.strftime("%Y-%m-%d")} to {student.session.end_year.strftime("%Y-%m-%d")}'
                    })
                
                existing_attendance = Attendance.objects.filter(
                    subject=subject,
                    date=attendance_date_obj
                ).first()
                
                if existing_attendance:
                    # Check if student already has attendance for this date
                    if AttendanceReport.objects.filter(
                        student=student,
                        attendance=existing_attendance
                    ).exists():
                        return JsonResponse({
                            'status': 'error',
                            'message': 'Attendance already marked for this student today'
                        })
                else:
                    # Create new attendance record
                    existing_attendance = Attendance.objects.create(
                        subject=subject,
                        session=student.session,
                        date=attendance_date_obj
                    )
                
                # Create attendance report
                attendance_report = AttendanceReport.objects.create(
                    student=student,
                    attendance=existing_attendance,
                    status=attendance_status
                )
                
                # Store the last successful attendance in session
                request.session['last_attendance'] = {
                    'student_name': f"{student.admin.first_name} {student.admin.last_name}",
                    'subject': subject.name,
                    'date': attendance_date,
                    'status': 'Present' if attendance_status else 'Absent',
                    'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                }
                
                session_str = student.session.start_year.strftime('%Y-%m-%d') + " to " + student.session.end_year.strftime('%Y-%m-%d')
                return JsonResponse({
                    'status': 'success',
                    'message': 'Attendance recorded successfully',
                    'student_name': f"{student.admin.first_name} {student.admin.last_name}",
                    'session': session_str,
                    'attendance_status': 'Present' if attendance_status else 'Absent'
                })
                
            except ValueError as e:
                print(f"[DEBUG] Date error: {str(e)}")
                return JsonResponse({
                    'status': 'error',
                    'message': f'Invalid date format: {attendance_date}. Expected format: YYYY-MM-DD'
                })
            except Exception as e:
                print(f"[DEBUG] Error recording attendance: {str(e)}")
                return JsonResponse({
                    'status': 'error',
                    'message': f'Error recording attendance: {str(e)}'
                })
            
        except Exception as e:
            import traceback
            print(f"Error in staff_take_attendance_by_qr: {str(e)}")
            print(f"Traceback: {traceback.format_exc()}")
            return JsonResponse({
                'status': 'error',
                'message': f'Error recording attendance: {str(e)}'
            })
    
    subjects = Subject.objects.filter(staff__admin=request.user)
    last_attendance = request.session.get('last_attendance', None)
    if last_attendance:
        del request.session['last_attendance']
        request.session.modified = True
    
    context = {
        'subjects': subjects,
        'page_title': 'Take Attendance by QR Code',
        'last_attendance': last_attendance
    }
    return render(request, 'staff_template/staff_take_attendance_by_qr.html', context)

@csrf_exempt
def get_students(request):
    subject_id = request.POST.get('subject')
    session_id = request.POST.get('session')

    if not subject_id or not session_id:
        return JsonResponse({
            'error': 'Both subject and session are required'
        }, status=400)

    try:
        subject = get_object_or_404(Subject, id=subject_id)
        session = get_object_or_404(Session, id=session_id)

        # Get students for this course and session
        students = Student.objects.filter(
            course_id=subject.course.id,
            session=session
        ).select_related('admin')  # Use select_related for better performance

        # Debug information
        print(f"Subject: {subject.name}, Course: {subject.course.name}")
        print(f"Session: {session}")
        print(f"Found {students.count()} students")

        student_data = []
        for student in students:
            data = {
                "id": student.id,
                "name": f"{student.admin.first_name} {student.admin.last_name}"
            }
            student_data.append(data)

        return JsonResponse(student_data, safe=False)

    except Subject.DoesNotExist:
        return JsonResponse({
            'error': f'Subject with id {subject_id} does not exist'
        }, status=404)
    except Session.DoesNotExist:
        return JsonResponse({
            'error': f'Session with id {session_id} does not exist'
        }, status=404)
    except Exception as e:
        print(f"Error in get_students: {str(e)}")
        return JsonResponse({
            'error': f'An error occurred: {str(e)}'
        }, status=500)
    
def staff_view_attendance(request):
    staff = get_object_or_404(Staff, admin=request.user)
    subjects = Subject.objects.filter(staff_id=staff)
    sessions = Session.objects.all()
    context = {
        'subjects': subjects,
        'sessions': sessions,
        'page_title': 'view Attendance'
    }

    return render(request, 'staff_template/staff_view_attendance.html', context)

@csrf_exempt
def get_student_attendance(request):
    attendance_date_id = request.POST.get('attendance_date_id')
    try:
        date = get_object_or_404(Attendance, id=attendance_date_id)
        attendance_data = AttendanceReport.objects.filter(attendance=date)
        student_data = []
        for attendance in attendance_data:
            data = {
                "id": attendance.id,  # This is the AttendanceReport ID
                "student_id": attendance.student.id,  # Add the actual student ID
                "name": attendance.student.admin.first_name + " " + attendance.student.admin.last_name,
                "status": attendance.status
            }
            student_data.append(data)
        return JsonResponse(json.dumps(student_data), content_type='application/json', safe=False)
    except Exception as e:
        return e

@csrf_exempt
def get_enrolled_students(request):
    subject_id = request.POST.get('subject')
    session_id = request.POST.get('session')
    
    if not subject_id or not session_id:
        return JsonResponse({
            'error': 'Both subject and session are required'
        }, status=400)
    
    try:
        subject = get_object_or_404(Subject, id=subject_id)
        session = get_object_or_404(Session, id=session_id)
        
        # Get all students enrolled in this course and session
        students = Student.objects.filter(
            course_id=subject.course.id,
            session=session
        ).select_related('admin')
        
        student_data = []
        for student in students:
            data = {
                "id": student.id,
                "name": f"{student.admin.first_name} {student.admin.last_name}"
            }
            student_data.append(data)
        
        return JsonResponse(student_data, safe=False)
    
    except Exception as e:
        print(f"Error in get_enrolled_students: {str(e)}")
        return JsonResponse({
            'error': f'An error occurred: {str(e)}'
        }, status=500)

def staff_apply_leave(request):
    form = LeaveReportStaffForm(request.POST or None)
    staff = get_object_or_404(Staff, admin_id=request.user.id)
    context = {
        'form': form,
        'leave_history': LeaveReportStaff.objects.filter(staff=staff),
        'page_title': 'Apply for Leave'
    }
    if request.method == 'POST':
        if form.is_valid():
            try:
                obj = form.save(commit=False)
                obj.staff = staff
                obj.save()
                messages.success(
                    request, "Application for leave has been submitted for review")
                return redirect(reverse('staff_apply_leave'))
            except Exception:
                messages.error(request, "Could not apply!")
        else:
            messages.error(request, "Form has errors!")
    return render(request, "staff_template/staff_apply_leave.html", context)


def staff_feedback(request):
    form = FeedbackStaffForm(request.POST or None)
    staff = get_object_or_404(Staff, admin_id=request.user.id)
    context = {
        'form': form,
        'feedbacks': FeedbackStaff.objects.filter(staff=staff),
        'page_title': 'Add Feedback'
    }
    if request.method == 'POST':
        if form.is_valid():
            try:
                obj = form.save(commit=False)
                obj.staff = staff
                obj.save()
                messages.success(request, "Feedback submitted for review")
                return redirect(reverse('staff_feedback'))
            except Exception:
                messages.error(request, "Could not Submit!")
        else:
            messages.error(request, "Form has errors!")
    return render(request, "staff_template/staff_feedback.html", context)


def staff_view_profile(request):
    staff = get_object_or_404(Staff, admin=request.user)
    form = StaffEditForm(request.POST or None, request.FILES or None,instance=staff)
    context = {'form': form, 'page_title': 'View/Update Profile'}
    if request.method == 'POST':
        try:
            if form.is_valid():
                first_name = form.cleaned_data.get('first_name')
                last_name = form.cleaned_data.get('last_name')
                password = form.cleaned_data.get('password') or None
                address = form.cleaned_data.get('address')
                gender = form.cleaned_data.get('gender')
                passport = request.FILES.get('profile_pic') or None
                admin = staff.admin
                if password != None:
                    admin.set_password(password)
                if passport != None:
                    fs = FileSystemStorage()
                    filename = fs.save(passport.name, passport)
                    passport_url = fs.url(filename)
                    admin.profile_pic = passport_url
                admin.first_name = first_name
                admin.last_name = last_name
                admin.address = address
                admin.gender = gender
                admin.save()
                staff.save()
                messages.success(request, "Profile Updated!")
                return redirect(reverse('staff_view_profile'))
            else:
                messages.error(request, "Invalid Data Provided")
                return render(request, "staff_template/staff_view_profile.html", context)
        except Exception as e:
            messages.error(
                request, "Error Occured While Updating Profile " + str(e))
            return render(request, "staff_template/staff_view_profile.html", context)

    return render(request, "staff_template/staff_view_profile.html", context)


@csrf_exempt
def staff_fcmtoken(request):
    token = request.POST.get('token')
    try:
        staff_user = get_object_or_404(CustomUser, id=request.user.id)
        staff_user.fcm_token = token
        staff_user.save()
        return HttpResponse("True")
    except Exception as e:
        return HttpResponse("False")


def staff_view_notification(request):
    staff = get_object_or_404(Staff, admin=request.user)
    notifications = NotificationStaff.objects.filter(staff=staff)
    context = {
        'notifications': notifications,
        'page_title': "View Notifications"
    }
    return render(request, "staff_template/staff_view_notification.html", context)


@csrf_exempt
def get_student_information(request):
    if request.method != "POST":
        return HttpResponse("Method Not Allowed", status=405)
    
    student_id = request.POST.get('student_id')
    print(f"[DEBUG] Received student_id: {student_id}")  # Debug log
    
    if not student_id:
        return JsonResponse({'error': 'Student ID is required'}, status=400)
    
    try:
        # Try to find student by student_code first
        student = None
        try:
            print(f"[DEBUG] Attempting to find student by student_code: {student_id}")
            student = Student.objects.select_related(
                'admin', 
                'session',
                'course'
            ).get(admin__student_code=student_id)
            print(f"[DEBUG] Found student by code: {student.admin.first_name} {student.admin.last_name}")
        except Student.DoesNotExist:
            # If not found by student_code, try finding by ID
            try:
                print(f"[DEBUG] Attempting to find student by ID: {student_id}")
                student = Student.objects.select_related(
                    'admin', 
                    'session',
                    'course'
                ).get(id=student_id)
                print(f"[DEBUG] Found student by ID: {student.admin.first_name} {student.admin.last_name}")
            except (Student.DoesNotExist, ValueError):
                print(f"[DEBUG] Student not found with code or ID: {student_id}")
                return JsonResponse({
                    'error': 'Student not found',
                    'details': 'No student found with the provided ID or code. Please verify the ID and try again.'
                }, status=404)
                
        if not student:
            return JsonResponse({
                'error': 'Student not found',
                'details': 'Unable to locate student record.'
            }, status=404)
            
        # Get student details
        if not student.session:
            print(f"[DEBUG] No session found for student: {student.admin.first_name} {student.admin.last_name}")
            # Try to find the latest session
            latest_session = Session.objects.order_by('-start_year').first()
            if latest_session:
                student.session = latest_session
                student.save()
                print(f"[DEBUG] Assigned latest session to student: {latest_session}")
            else:
                return JsonResponse({
                    'error': 'No session assigned',
                    'details': 'This student does not have a session assigned.'
                }, status=400)
                
        session_str = student.session.start_year.strftime('%Y-%m-%d') + " to " + student.session.end_year.strftime('%Y-%m-%d')
        
        # Get profile picture URL
        profile_pic_url = None
        if hasattr(student.admin, 'profile_pic') and student.admin.profile_pic:
            profile_pic_url = student.admin.profile_pic.url
            
        # Get the current staff member
        staff = Staff.objects.get(admin=request.user)
        
        # Get subjects that this student can attend and are taught by the current staff
        registered_subjects = []
        
        # Try to find subjects through the StudentSubject model if it exists
        try:
            # Check if the StudentSubject model is being used for subject registration
            student_subjects = StudentSubject.objects.filter(student=student)
            if student_subjects.exists():
                # Filter for subjects taught by this staff member
                for subject_entry in student_subjects:
                    if subject_entry.subject.staff_id == staff.id:
                        registered_subjects.append({
                            'id': subject_entry.subject.id,
                            'name': subject_entry.subject.name
                        })
        except:
            # If StudentSubject model doesn't exist or no records found, we'll continue below
            pass
            
        # If no registered subjects found through StudentSubject model,
        # find all subjects in the student's course taught by this staff member
        if not registered_subjects:
            course_subjects = Subject.objects.filter(
                course=student.course,
                staff=staff
            )
            for subject in course_subjects:
                registered_subjects.append({
                    'id': subject.id,
                    'name': subject.name
                })
        
        # Additional student information from manage_student.html
        student_data = {
            'name': f"{student.admin.first_name} {student.admin.last_name}",
            'first_name': student.admin.first_name,
            'last_name': student.admin.last_name,
            'email': student.admin.email,
            'gender': student.admin.gender,
            'session': session_str,
            'session_start_date': student.session.start_year.strftime('%Y-%m-%d'),
            'session_end_date': student.session.end_year.strftime('%Y-%m-%d'),
            'course': student.course.name if student.course else None,
            'session_id': student.session.id if student.session else None,
            'session_years': f"{student.session.start_year.year} - {student.session.end_year.year}" if student.session else None,
            'student_code': student.admin.student_code if hasattr(student.admin, 'student_code') else None,
            'qr_code': student.admin.qr_code.url.replace('/media/', '') if hasattr(student.admin, 'qr_code') and student.admin.qr_code else None,
            'profile_pic': profile_pic_url.replace('/media/', '') if profile_pic_url else None,  # Add profile picture URL
            'registered_subjects': registered_subjects  # Add the registered subjects data
        }
        
        print(f"[DEBUG] Returning student data: {student_data}")
        return JsonResponse(student_data)
    except Exception as e:
        print(f"[DEBUG] Error in get_student_information: {str(e)}")
        import traceback
        print(f"[DEBUG] Full traceback: {traceback.format_exc()}")
        return JsonResponse({
            'error': 'An error occurred while looking up the student',
            'details': str(e)
        }, status=500)

@csrf_exempt
def get_recent_attendance_records(request):
    """
    Get recent attendance records for staff dashboard with precise timestamps
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Invalid request method'}, status=405)
    
    try:
        # Get the staff member
        staff = Staff.objects.get(admin=request.user)
        
        # Get subjects taught by this staff
        subjects = Subject.objects.filter(staff=staff)
        
        # Get recent attendance reports for these subjects (limit to 20)
        attendance_records = Attendance.objects.filter(
            subject__in=subjects
        ).order_by('-created_at')[:20]
        
        records = []
        for attendance in attendance_records:
            # Get all reports for this attendance record
            reports = AttendanceReport.objects.filter(attendance=attendance)
            
            # Count present and absent students
            present_count = reports.filter(status=True).count()
            absent_count = reports.filter(status=False).count()
            total_count = present_count + absent_count
            
            if total_count > 0:
                present_percentage = round((present_count / total_count) * 100)
            else:
                present_percentage = 0
                
            # Format the created_at timestamp with minutes and seconds
            created_at = attendance.created_at
            time_str = created_at.strftime('%Y-%m-%d %H:%M:%S')
            
            # Calculate how long ago the record was created
            now = timezone.now()
            time_diff = now - created_at
            
            # Format time difference in a human-readable way
            if time_diff.days > 0:
                time_ago = f"{time_diff.days} day{'s' if time_diff.days > 1 else ''} ago"
            elif time_diff.seconds // 3600 > 0:
                hours = time_diff.seconds // 3600
                time_ago = f"{hours} hour{'s' if hours > 1 else ''} ago"
            elif time_diff.seconds // 60 > 0:
                minutes = time_diff.seconds // 60
                time_ago = f"{minutes} minute{'s' if minutes > 1 else ''} ago"
            else:
                time_ago = f"{time_diff.seconds} second{'s' if time_diff.seconds != 1 else ''} ago"
            
            # Get student details for this attendance record
            student_details = []
            for report in reports:
                student = report.student
                student_details.append({
                    'id': student.id,
                    'name': f"{student.admin.first_name} {student.admin.last_name}",
                    'status': report.status
                })
                
            records.append({
                'id': attendance.id,
                'subject': attendance.subject.name,
                'date': attendance.date.strftime('%Y-%m-%d'),
                'created_at': time_str,
                'time_ago': time_ago,
                'present_count': present_count,
                'absent_count': absent_count,
                'total_count': total_count,
                'present_percentage': present_percentage,
                'students': student_details
            })
        
        return JsonResponse({'status': 'success', 'records': records})
    
    except Exception as e:
        import traceback
        print(f"Error in get_recent_attendance_records: {str(e)}")
        print(traceback.format_exc())
        return JsonResponse({
            'status': 'error',
            'message': f'An error occurred: {str(e)}'
        })

@csrf_exempt
def export_attendance_to_excel(request):
    """
    Export attendance data to Excel for staff members
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Invalid request method'}, status=405)
    
    try:
        # Get the staff member
        staff = Staff.objects.get(admin=request.user)
        
        # Get parameters from request
        subject_id = request.POST.get('subject_id')
        session_id = request.POST.get('session_id')
        
        if not subject_id or not session_id:
            return JsonResponse({
                'status': 'error',
                'message': 'Subject and session are required'
            })
        
        # Get the subject and session
        subject = get_object_or_404(Subject, id=subject_id)
        session = get_object_or_404(Session, id=session_id)
        
        # Verify that the subject belongs to this staff
        if subject.staff_id != staff.id:
            return JsonResponse({
                'status': 'error',
                'message': 'You are not authorized to access this subject'
            })
        
        # Get all students in this course and session
        students = Student.objects.filter(
            course=subject.course,
            session=session
        ).select_related('admin')
        
        # Get all attendance records for this subject and session
        attendance_records = Attendance.objects.filter(
            subject=subject,
            session=session
        ).order_by('date')
        
        # Create a new workbook
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from io import BytesIO
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{subject.name} Attendance"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="344054", end_color="344054", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        present_fill = PatternFill(start_color="DDFFDD", end_color="DDFFDD", fill_type="solid")
        absent_fill = PatternFill(start_color="FFDDDD", end_color="FFDDDD", fill_type="solid")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add title
        ws.merge_cells('A1:E1')
        title_cell = ws['A1']
        title_cell.value = f"Attendance Report - {subject.name} ({session})"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")
        
        # Add metadata
        ws['A2'] = "Course:"
        ws['B2'] = subject.course.name
        ws['A3'] = "Subject:"
        ws['B3'] = subject.name
        ws['A4'] = "Session:"
        ws['B4'] = str(session)
        ws['A5'] = "Staff:"
        ws['B5'] = f"{staff.admin.first_name} {staff.admin.last_name}"
        ws['A6'] = "Generated on:"
        ws['B6'] = timezone.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # Style metadata
        for row in range(2, 7):
            ws[f'A{row}'].font = Font(bold=True)
        
        # Add header row for attendance dates
        start_row = 8
        ws.cell(row=start_row, column=1, value="No.").font = header_font
        ws.cell(row=start_row, column=1).fill = header_fill
        ws.cell(row=start_row, column=1).alignment = header_alignment
        
        ws.cell(row=start_row, column=2, value="Student Name").font = header_font
        ws.cell(row=start_row, column=2).fill = header_fill
        ws.cell(row=start_row, column=2).alignment = header_alignment
        
        # Add attendance dates in header
        col_index = 3
        for attendance in attendance_records:
            date_formatted = attendance.date.strftime("%b %d, %Y")
            ws.cell(row=start_row, column=col_index, value=date_formatted).font = header_font
            ws.cell(row=start_row, column=col_index).fill = header_fill
            ws.cell(row=start_row, column=col_index).alignment = header_alignment
            col_index += 1
        
        # Add "Present" and "Absent" summary columns
        ws.cell(row=start_row, column=col_index, value="Present").font = header_font
        ws.cell(row=start_row, column=col_index).fill = header_fill
        ws.cell(row=start_row, column=col_index).alignment = header_alignment
        
        ws.cell(row=start_row, column=col_index+1, value="Absent").font = header_font
        ws.cell(row=start_row, column=col_index+1).fill = header_fill
        ws.cell(row=start_row, column=col_index+1).alignment = header_alignment
        
        ws.cell(row=start_row, column=col_index+2, value="Percentage").font = header_font
        ws.cell(row=start_row, column=col_index+2).fill = header_fill
        ws.cell(row=start_row, column=col_index+2).alignment = header_alignment
        
        # Add student data rows
        row_index = start_row + 1
        student_num = 1
        
        for student in students:
            # Add student name
            ws.cell(row=row_index, column=1, value=student_num)
            ws.cell(row=row_index, column=2, value=f"{student.admin.first_name} {student.admin.last_name}")
            
            # Add attendance status for each date
            col_index = 3
            present_count = 0
            absent_count = 0
            
            for attendance in attendance_records:
                # Check if there's a report for this student
                try:
                    report = AttendanceReport.objects.get(
                        student=student,
                        attendance=attendance
                    )
                    if report.status:  # Present
                        status = "P"
                        present_count += 1
                        ws.cell(row=row_index, column=col_index).fill = present_fill
                    else:  # Absent
                        status = "A"
                        absent_count += 1
                        ws.cell(row=row_index, column=col_index).fill = absent_fill
                except AttendanceReport.DoesNotExist:
                    status = "A"  # Default to absent if no record
                    absent_count += 1
                    ws.cell(row=row_index, column=col_index).fill = absent_fill
                
                ws.cell(row=row_index, column=col_index, value=status)
                ws.cell(row=row_index, column=col_index).alignment = Alignment(horizontal="center")
                col_index += 1
            
            # Add summary columns
            total_days = present_count + absent_count
            percentage = round((present_count / total_days) * 100) if total_days > 0 else 0
            
            ws.cell(row=row_index, column=col_index, value=present_count)
            ws.cell(row=row_index, column=col_index).alignment = Alignment(horizontal="center")
            
            ws.cell(row=row_index, column=col_index+1, value=absent_count)
            ws.cell(row=row_index, column=col_index+1).alignment = Alignment(horizontal="center")
            
            ws.cell(row=row_index, column=col_index+2, value=f"{percentage}%")
            ws.cell(row=row_index, column=col_index+2).alignment = Alignment(horizontal="center")
            
            row_index += 1
            student_num += 1
        
        # Add borders to all cells
        for row in ws.iter_rows(min_row=start_row, max_row=row_index-1, 
                              min_col=1, max_col=col_index+2):
            for cell in row:
                cell.border = border
        
        # Adjust column widths
        for col in range(1, col_index+3):
            column_letter = get_column_letter(col)
            if col == 2:  # Student name column
                ws.column_dimensions[column_letter].width = 25
            else:
                ws.column_dimensions[column_letter].width = 12
        
        # Freeze panes to make the header row and student names stay visible
        ws.freeze_panes = 'C9'
        
        # Save workbook to BytesIO object
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename
        filename = f"{subject.name}_{session}_attendance_report.xlsx"
        
        # Prepare response
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
    
    except Exception as e:
        import traceback
        print(f"Error in export_attendance_to_excel: {str(e)}")
        print(traceback.format_exc())
        return JsonResponse({
            'status': 'error',
            'message': f'An error occurred: {str(e)}'
        })
