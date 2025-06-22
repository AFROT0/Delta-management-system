# Standard library imports
import json
import logging
import random
import string
from datetime import datetime
from io import BytesIO
import os

# Third-party imports
import requests
from django.contrib import messages
from django.core.files.base import ContentFile
from django.core.files.storage import FileSystemStorage
from django.http import HttpResponse, JsonResponse, HttpResponseRedirect
from django.shortcuts import get_object_or_404, redirect, render
from django.templatetags.static import static
from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt
from django.views.generic import UpdateView
from django.db import transaction
try:
   import qrcode
except ImportError:
    qrcode = None

# Local application imports
from .forms import *
from .models import (
    Admin, Attendance, AttendanceReport, Course, CustomUser,
    FeedbackStaff, FeedbackStudent, LeaveReportStaff,
    LeaveReportStudent, NotificationStaff, NotificationStudent,
    Session, Staff, Student, Subject, StudentSubject
)
from django.core import serializers
from django.contrib.auth.models import Group
from django.contrib.auth.decorators import login_required


def admin_home(request):
    total_staff = Staff.objects.all().count()
    total_students = Student.objects.all().count()
    subjects = Subject.objects.all()
    total_subject = subjects.count()
    total_course = Course.objects.all().count()
    
    # Attendance per subject
    subject_list = []
    attendance_list = []
    for subject in subjects:
        attendance_count = Attendance.objects.filter(subject=subject).count()
        subject_list.append(subject.name[:7])
        attendance_list.append(attendance_count)

    # Students per course
    course_all = Course.objects.all()
    course_name_list = []
    student_count_list_in_course = []
    for course in course_all:
        students = Student.objects.filter(course_id=course.id).count()
        course_name_list.append(course.name)
        student_count_list_in_course.append(students)
    
    # Students per subject
    subject_all = Subject.objects.all()
    subject_list = []
    student_count_list_in_subject = []
    for subject in subject_all:
        course = Course.objects.get(id=subject.course.id)
        student_count = Student.objects.filter(course_id=course.id).count()
        subject_list.append(subject.name)
        student_count_list_in_subject.append(student_count)

    # Student attendance statistics
    student_attendance_present_list = []
    student_attendance_leave_list = []
    student_name_list = []
    students = Student.objects.all()
    for student in students:
        attendance = AttendanceReport.objects.filter(student_id=student.id, status=True).count()
        absent = AttendanceReport.objects.filter(student_id=student.id, status=False).count()
        leave = LeaveReportStudent.objects.filter(student_id=student.id, status=1).count()
        student_attendance_present_list.append(attendance)
        student_attendance_leave_list.append(leave + absent)
        student_name_list.append(f"{student.admin.first_name} {student.admin.last_name}")

    context = {
        'page_title': "Administrative Dashboard",
        'total_students': total_students,
        'total_staff': total_staff,
        'total_course': total_course,
        'total_subject': total_subject,
        'subject_list': subject_list,
        'attendance_list': attendance_list,
        'student_attendance_present_list': student_attendance_present_list,
        'student_attendance_leave_list': student_attendance_leave_list,
        'student_name_list': student_name_list,
        'student_count_list_in_subject': student_count_list_in_subject,
        'student_count_list_in_course': student_count_list_in_course,
        'course_name_list': course_name_list,
    }
    return render(request, 'hod_template/home_content.html', context)


def add_staff(request):
    form = StaffForm(request.POST or None, request.FILES or None)
    context = {'form': form, 'page_title': 'Add Staff'}
    if request.method == 'POST':
        if form.is_valid():
            first_name = form.cleaned_data.get('first_name')
            last_name = form.cleaned_data.get('last_name')
            address = form.cleaned_data.get('address')
            email = form.cleaned_data.get('email')
            gender = form.cleaned_data.get('gender')
            password = form.cleaned_data.get('password')
            course = form.cleaned_data.get('course')
            passport = request.FILES.get('profile_pic')
            
            try:
                user = CustomUser.objects.create_user(
                    email=email, password=password, user_type=2, first_name=first_name, last_name=last_name)
                user.gender = gender
                user.address = address
                user.staff.course = course
                
                if passport:
                    fs = FileSystemStorage()
                    filename = fs.save(passport.name, passport)
                    passport_url = fs.url(filename)
                    user.profile_pic = passport_url
                
                user.save()
                messages.success(request, "Staff added successfully!")
                return redirect(reverse('add_staff'))
            except Exception as e:
                messages.error(request, f"Could not add staff: {str(e)}")
        else:
            messages.error(request, "Please fill all required fields")

    return render(request, 'hod_template/add_staff_template.html', context)


def add_student(request):
    if request.method == 'POST':
        form = StudentForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                first_name = form.cleaned_data.get('first_name')
                last_name = form.cleaned_data.get('last_name')
                address = form.cleaned_data.get('address')
                email = form.cleaned_data.get('email')
                password = form.cleaned_data.get('password')
                course = form.cleaned_data.get('course')
                gender = form.cleaned_data.get('gender')
                session = form.cleaned_data.get('session')
                
                # Get subject ID from POST data (not in form)
                subject_id = request.POST.get('subject')
                
                # Generate a unique student code
                student_code = generate_unique_student_code()
                
                # Create user first
                user = CustomUser.objects.create_user(
                    email=email, password=password, user_type=3,
                    first_name=first_name, last_name=last_name)
                user.gender = gender
                user.address = address
                user.student.course = course
                user.student.session = session
                user.student_code = student_code
                
                # Handle profile picture if provided
                if 'profile_pic' in request.FILES:
                    profile_pic = request.FILES['profile_pic']
                    fs = FileSystemStorage()
                    filename = fs.save(profile_pic.name, profile_pic)
                    user.profile_pic = fs.url(filename)
                
                # Save user to get the ID
                user.save()
                
                # Instead of creating a new Student, get the one created by the signal
                student = Student.objects.get(admin=user)
                student.course = course
                student.session = session
                student.save()
                
                # Generate QR code for the student
                try:
                    if qrcode is not None:
                        # Create QR code with student data - only include student_code
                        student_data = {
                            'student_code': student_code
                        }
                        
                        # Create QR code
                        qr = qrcode.QRCode(
                            version=1,
                            error_correction=qrcode.constants.ERROR_CORRECT_L,
                            box_size=10,
                            border=4,
                        )
                        qr.add_data(json.dumps(student_data))
                        qr.make(fit=True)
                        qr_image = qr.make_image(fill_color="black", back_color="white")
                        
                        # Save QR code
                        buffer = BytesIO()
                        qr_image.save(buffer, format='PNG')
                        buffer.seek(0)
                        
                        # Create a unique filename for the QR code
                        filename = f'qr_code_{student_code}.png'
                        file_path = os.path.join('qr_codes', filename)
                        
                        # Use FileSystemStorage to save the file
                        fs = FileSystemStorage()
                        saved_path = fs.save(file_path, ContentFile(buffer.getvalue()))
                        
                        # Update user's qr_code field with the saved file path
                        user.qr_code = saved_path
                        user.save()
                    else:
                        messages.warning(request, "QR code generation is not available. Please install qrcode package.")
                except Exception as qr_error:
                    messages.warning(request, f"Could not generate QR code: {str(qr_error)}")
                
                # Create StudentSubject record if subject was selected
                if subject_id:
                    try:
                        subject = Subject.objects.get(id=subject_id)
                        StudentSubject.objects.create(student=student, subject=subject)
                    except Subject.DoesNotExist:
                        messages.error(request, "Selected subject does not exist")
                    except Exception as e:
                        messages.error(request, f"Error assigning subject: {str(e)}")
                
                messages.success(request, "Successfully Added")
                return redirect(reverse('add_student'))
            except Exception as e:
                messages.error(request, "Could Not Add: " + str(e))
        else:
            messages.error(request, "Form validation failed")
    else:
        form = StudentForm()
    return render(request, 'hod_template/add_student_template.html', {"form": form, "page_title": "Add Student"})

def generate_unique_student_code():
    import random
    import string
    
    while True:
        # Generate a 6-character code with only letters and numbers
        characters = string.ascii_letters + string.digits
        code = ''.join(random.choice(characters) for _ in range(6))
        
        # Check if code already exists
        if not CustomUser.objects.filter(student_code=code).exists():
            return code


def add_course(request):
    form = CourseForm(request.POST or None)
    context = {
        'form': form,
        'page_title': 'Add Course'
    }
    if request.method == 'POST':
        if form.is_valid():
            name = form.cleaned_data.get('name')
            # Check if course with the same name already exists
            if Course.objects.filter(name__iexact=name).exists():
                messages.error(request, "Course with this name already exists")
            else:
                try:
                    course = Course()
                    course.name = name
                    course.save()
                    messages.success(request, "Successfully Added")
                    return redirect(reverse('add_course'))
                except Exception as e:
                    messages.error(request, f"Could Not Add: {str(e)}")
        else:
            messages.error(request, "Could Not Add: Form is invalid")
    return render(request, 'hod_template/add_course_template.html', context)




def add_subject(request):
    form = SubjectForm(request.POST or None)
    context = {
        'form': form,
        'page_title': 'Add Subject'
    }
    if request.method == 'POST':
        if form.is_valid():
            name = form.cleaned_data.get('name')
            course = form.cleaned_data.get('course')
            staff = form.cleaned_data.get('staff')
            
            # Check if subject with the same name already exists in the same course
            if Subject.objects.filter(name__iexact=name, course=course).exists():
                messages.error(request, "Subject with this name already exists in the selected course")
            else:
                try:
                    subject = Subject()
                    subject.name = name
                    subject.staff = staff
                    subject.course = course
                    subject.save()
                    messages.success(request, "Successfully Added")
                    return redirect(reverse('add_subject'))
                except Exception as e:
                    messages.error(request, f"Could Not Add: {str(e)}")
        else:
            messages.error(request, "Please fill all required fields correctly")

    return render(request, 'hod_template/add_subject_template.html', context)


def manage_staff(request):
    allStaff = CustomUser.objects.filter(user_type=2)
    context = {
        'allStaff': allStaff,
        'page_title': 'Manage Staff'
    }
    return render(request, "hod_template/manage_staff.html", context)


def manage_student(request):
    students = CustomUser.objects.filter(user_type=3)
    courses = Course.objects.all()
    sessions = Session.objects.all()
    context = {
        'students': students,
        'courses': courses,
        'sessions': sessions,
        'page_title': 'Manage Students'
    }
    return render(request, "hod_template/manage_student.html", context)


def download_student_template(request):
    """
    Generate and provide Excel template for student import
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    from django.http import HttpResponse
    from django.utils import timezone
    
    # Create workbook and select active sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student Import Template"
    
    # Define styles
    header_font = Font(name='Arial', bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    # Add instructions
    ws.merge_cells('A1:H1')
    ws.cell(row=1, column=1).value = "Instructions for Student Import"
    ws.cell(row=1, column=1).font = Font(name='Arial', bold=True, size=14)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')
    
    # Add multiple instruction rows
    instructions = [
        "1. Fill in the student details according to the headers below.",
        "2. Fields marked with * are required.",
        "3. For Gender, use 'M' for Male and 'F' for Female.",
        "4. Course must match an existing course name in the system.",
        "5. Session will be selected from the dropdown when importing.",
        "6. Subject is optional and should match existing subject name(s).",
        "7. Student Code is auto-generated if left blank."
    ]
    
    for i, instruction in enumerate(instructions, 2):
        ws.merge_cells(f'A{i}:H{i}')
        ws.cell(row=i, column=1).value = instruction
        ws.cell(row=i, column=1).font = Font(name='Arial', italic=True)
    
    # Header row (row 10)
    headers = [
        'First Name*', 
        'Last Name*', 
        'Email*', 
        'Gender*', 
        'Password*', 
        'Address',
        'Course*',
        'Subject'
    ]
    
    row_num = 10
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Add example data
    example_data = [
        'John',
        'Doe',
        'john.doe@example.com',
        'M',
        'password123',
        '123 Main St, City',
        'Computer Science',
        'Programming'
    ]
    
    row_num = 11
    for col_num, value in enumerate(example_data, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = value
        cell.border = thin_border
    
    # Add 20 empty rows for data entry
    for row in range(12, 32):
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = thin_border
    
    # Adjust column widths - fixed to handle merged cells
    column_widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                # Skip cells that are part of a merged range
                if isinstance(cell, openpyxl.cell.cell.MergedCell):
                    continue
                
                col_letter = openpyxl.utils.get_column_letter(cell.column)
                current_width = column_widths.get(col_letter, 0)
                column_widths[col_letter] = max(current_width, len(str(cell.value)) + 2)
    
    # Apply the calculated widths
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Create the HttpResponse with Excel content type
    filename = f"Student_Import_Template_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


def import_students_from_excel(request):
    """
    Import students from Excel file
    """
    from django.contrib import messages
    import openpyxl
    import random
    import string
    
    if request.method != 'POST':
        return redirect(reverse('manage_student'))
    
    excel_file = request.FILES.get('excel_file')
    if not excel_file:
        messages.error(request, "Please select an Excel file")
        return redirect(reverse('manage_student'))
    
    session_id = request.POST.get('session_id')
    if not session_id:
        messages.error(request, "Please select a session")
        return redirect(reverse('manage_student'))
    
    try:
        session = Session.objects.get(id=session_id)
    except Session.DoesNotExist:
        messages.error(request, "Selected session does not exist")
        return redirect(reverse('manage_student'))
    
    if not excel_file.name.endswith(('.xlsx', '.xls')):
        messages.error(request, "File must be an Excel file (.xlsx or .xls)")
        return redirect(reverse('manage_student'))
    
    try:
        # Load workbook and select active sheet
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        
        # Column indices (1-based)
        first_name_col, last_name_col, email_col = 1, 2, 3
        gender_col, password_col, address_col = 4, 5, 6
        course_col, subject_col = 7, 8
        
        successful_imports = 0
        failed_imports = 0
        error_messages = []
        
        # Start from row 11 (skipping headers and example)
        for row in range(11, ws.max_row + 1):
            # Skip rows with no data
            if not ws.cell(row=row, column=first_name_col).value:
                continue
                
            try:
                # Extract data from the row
                first_name = ws.cell(row=row, column=first_name_col).value
                last_name = ws.cell(row=row, column=last_name_col).value
                email = ws.cell(row=row, column=email_col).value
                gender = ws.cell(row=row, column=gender_col).value
                password = ws.cell(row=row, column=password_col).value
                address = ws.cell(row=row, column=address_col).value or ""
                course_name = ws.cell(row=row, column=course_col).value
                subject_name = ws.cell(row=row, column=subject_col).value
                
                # Validate required fields
                if not all([first_name, last_name, email, gender, password, course_name]):
                    error_messages.append(f"Row {row}: Missing required fields")
                    failed_imports += 1
                    continue
                
                # Validate email format
                if '@' not in email:
                    error_messages.append(f"Row {row}: Invalid email format")
                    failed_imports += 1
                    continue
                
                # Validate gender
                if gender not in ['M', 'F']:
                    error_messages.append(f"Row {row}: Gender must be 'M' or 'F'")
                    failed_imports += 1
                    continue
                
                # Check if email already exists
                if CustomUser.objects.filter(email=email).exists():
                    error_messages.append(f"Row {row}: Email {email} already exists")
                    failed_imports += 1
                    continue
                
                # Find the course
                try:
                    course = Course.objects.get(name=course_name)
                except Course.DoesNotExist:
                    error_messages.append(f"Row {row}: Course '{course_name}' does not exist")
                    failed_imports += 1
                    continue
                
                # Find the subject (optional)
                subject = None
                if subject_name:
                    try:
                        subject = Subject.objects.get(name=subject_name, course=course)
                    except Subject.DoesNotExist:
                        error_messages.append(f"Row {row}: Subject '{subject_name}' does not exist for course '{course_name}'")
                        failed_imports += 1
                        continue
                
                # Generate student code using the existing function
                student_code = generate_unique_student_code()
                
                # Create the user and student
                user = CustomUser.objects.create_user(
                    email=email,
                    password=password,
                    first_name=first_name,
                    last_name=last_name,
                    user_type=3,  # 3 for student
                    gender=gender,
                    address=address,
                    student_code=student_code
                )
                
                # Update student data
                student = Student.objects.get(admin=user)
                student.course = course
                student.session = session
                student.save()
                
                # Create StudentSubject record if subject was selected
                if subject:
                    try:
                        StudentSubject.objects.create(student=student, subject=subject)
                    except Exception as e:
                        # Non-critical error, continue anyway
                        error_messages.append(f"Row {row}: Error adding subject: {str(e)}")
                
                successful_imports += 1
                
            except Exception as e:
                error_messages.append(f"Row {row}: {str(e)}")
                failed_imports += 1
        
        # Show summary message
        if successful_imports > 0:
            messages.success(request, f"Successfully imported {successful_imports} student(s)")
        
        if failed_imports > 0:
            messages.warning(request, f"Failed to import {failed_imports} student(s). See details below.")
            for error in error_messages[:10]:  # Show first 10 errors
                messages.error(request, error)
            
            if len(error_messages) > 10:
                messages.error(request, f"...and {len(error_messages) - 10} more errors")
        
    except Exception as e:
        messages.error(request, f"Error processing file: {str(e)}")
    
    return redirect(reverse('manage_student'))


def manage_course(request):
    courses = Course.objects.all()
    context = {
        'courses': courses,
        'page_title': 'Manage Courses'
    }
    return render(request, "hod_template/manage_course.html", context)


def manage_subject(request):
    subjects = Subject.objects.all()
    context = {
        'subjects': subjects,
        'page_title': 'Manage Subjects'
    }
    return render(request, "hod_template/manage_subject.html", context)


def edit_staff(request, staff_id):
    staff = get_object_or_404(Staff, id=staff_id)
    form = StaffForm(request.POST or None, instance=staff)
    context = {
        'form': form,
        'staff_id': staff_id,
        'page_title': 'Edit Staff'
    }
    if request.method == 'POST':
        if form.is_valid():
            first_name = form.cleaned_data.get('first_name')
            last_name = form.cleaned_data.get('last_name')
            address = form.cleaned_data.get('address')
            username = form.cleaned_data.get('username')
            email = form.cleaned_data.get('email')
            gender = form.cleaned_data.get('gender')
            password = form.cleaned_data.get('password') or None
            course = form.cleaned_data.get('course')
            passport = request.FILES.get('profile_pic') or None
            try:
                user = CustomUser.objects.get(id=staff.admin.id)
                user.username = username
                user.email = email
                if password != None:
                    user.set_password(password)
                if passport != None:
                    fs = FileSystemStorage()
                    filename = fs.save(passport.name, passport)
                    passport_url = fs.url(filename)
                    user.profile_pic = passport_url
                user.first_name = first_name
                user.last_name = last_name
                user.gender = gender
                user.address = address
                staff.course = course
                user.save()
                staff.save()
                messages.success(request, "Successfully Updated")
                return redirect(reverse('edit_staff', args=[staff_id]))
            except Exception as e:
                messages.error(request, "Could Not Update " + str(e))
        else:
            messages.error(request, "Please fil form properly")
    else:
        user = CustomUser.objects.get(id=staff_id)
        staff = Staff.objects.get(id=user.id)
        return render(request, "hod_template/edit_staff_template.html", context)


def edit_student(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    form = StudentEditForm(request.POST or None, request.FILES or None, instance=student)
    context = {
        'form': form,
        'student_id': student_id,
        'page_title': 'Edit Student'
    }
    if request.method == 'POST':
        if form.is_valid():
            first_name = form.cleaned_data.get('first_name')
            last_name = form.cleaned_data.get('last_name')
            address = form.cleaned_data.get('address')
            email = form.cleaned_data.get('email')
            gender = form.cleaned_data.get('gender')
            password = form.cleaned_data.get('password') or None
            student_code = form.cleaned_data.get('student_code')
            course = form.cleaned_data.get('course')
            session = form.cleaned_data.get('session')
            passport = request.FILES.get('profile_pic') or None
            
            # Get subject ID from POST data
            subject_id = request.POST.get('subject')
            
            try:
                user = CustomUser.objects.get(id=student.admin.id)
                if passport != None:
                    fs = FileSystemStorage()
                    filename = fs.save(passport.name, passport)
                    passport_url = fs.url(filename)
                    user.profile_pic = passport_url
                user.email = email
                if password != None:
                    user.set_password(password)
                user.first_name = first_name
                user.last_name = last_name
                user.gender = gender
                user.address = address
                
                # Update student fields
                student.course = course
                student.session = session
                
                # Update student code if provided and valid (validation happens in form)
                if student_code:
                    user.student_code = student_code
                user.save()
                student.save()
                
                # Handle subject - first remove existing if any
                if subject_id:
                    # Update or create subject association
                    try:
                        subject = Subject.objects.get(id=subject_id)
                        # Remove existing subject associations first
                        StudentSubject.objects.filter(student=student).delete()
                        # Create new association
                        StudentSubject.objects.create(student=student, subject=subject)
                    except Subject.DoesNotExist:
                        messages.error(request, "Selected subject does not exist")
                    except Exception as e:
                        messages.error(request, f"Error assigning subject: {str(e)}")
                
                messages.success(request, "Successfully Updated")
                return redirect(reverse('edit_student', args=[student_id]))
            except Exception as e:
                messages.error(request, "Could Not Update " + str(e))
        else:
            messages.error(request, "Form validation failed")
    
    # Get student subjects to display in template
    student_subjects = StudentSubject.objects.filter(student=student)
    context['student_subjects'] = student_subjects
    
    return render(request, "hod_template/edit_student.html", context)


def edit_course(request, course_id):
    instance = get_object_or_404(Course, id=course_id)
    form = CourseForm(request.POST or None, instance=instance)
    context = {
        'form': form,
        'course_id': course_id,
        'page_title': 'Edit Course'
    }
    if request.method == 'POST':
        if form.is_valid():
            name = form.cleaned_data.get('name')
            try:
                course = Course.objects.get(id=course_id)
                course.name = name
                course.save()
                messages.success(request, "Successfully Updated")
            except:
                messages.error(request, "Could Not Update")
        else:
            messages.error(request, "Could Not Update")

    return render(request, 'hod_template/edit_course_template.html', context)


def edit_subject(request, subject_id):
    instance = get_object_or_404(Subject, id=subject_id)
    form = SubjectForm(request.POST or None, instance=instance)
    context = {
        'form': form,
        'subject_id': subject_id,
        'page_title': 'Edit Subject'
    }
    if request.method == 'POST':
        if form.is_valid():
            name = form.cleaned_data.get('name')
            course = form.cleaned_data.get('course')
            staff = form.cleaned_data.get('staff')
            try:
                subject = Subject.objects.get(id=subject_id)
                subject.name = name
                subject.staff = staff
                subject.course = course
                subject.save()
                messages.success(request, "Successfully Updated")
                return redirect(reverse('edit_subject', args=[subject_id]))
            except Exception as e:
                messages.error(request, "Could Not Add " + str(e))
        else:
            messages.error(request, "Fill Form Properly")
    return render(request, 'hod_template/edit_subject_template.html', context)


def add_session(request):
    form = SessionForm(request.POST or None)
    context = {'form': form, 'page_title': 'Add Session'}
    
    if request.method == 'POST':
        if form.is_valid():
            start_year = form.cleaned_data.get('start_year')
            end_year = form.cleaned_data.get('end_year')
            
            # Check if start date equals end date
            if start_year == end_year:
                messages.error(request, 'Start date cannot be the same as end date')
                return render(request, "hod_template/add_session_template.html", context)
            
            # Check if start date is after end date
            if start_year > end_year:
                messages.error(request, 'Start date cannot be after end date')
                return render(request, "hod_template/add_session_template.html", context)
            
            # Check if session with same dates already exists
            if Session.objects.filter(start_year=start_year, end_year=end_year).exists():
                messages.error(request, 'Session with these dates already exists')
                return render(request, "hod_template/add_session_template.html", context)
            
            try:
                form.save()
                messages.success(request, "Session was created successfully")
                return redirect(reverse('add_session'))
            except Exception as e:
                messages.error(request, f'Failed to add session: {str(e)}')
        else:
            # This shows when form validation fails
            messages.error(request, 'This session already exists')
    
    return render(request, "hod_template/add_session_template.html", context)

def manage_session(request):
    sessions = Session.objects.all()
    context = {'sessions': sessions, 'page_title': 'Manage Sessions'}
    return render(request, "hod_template/manage_session.html", context)


def edit_session(request, session_id):
    instance = get_object_or_404(Session, id=session_id)
    form = SessionForm(request.POST or None, instance=instance)
    context = {'form': form, 'session_id': session_id,
               'page_title': 'Edit Session'}
    if request.method == 'POST':
        if form.is_valid():
            try:
                form.save()
                messages.success(request, "Session Updated")
                return redirect(reverse('edit_session', args=[session_id]))
            except Exception as e:
                messages.error(
                    request, "Session Could Not Be Updated " + str(e))
                return render(request, "hod_template/edit_session_template.html", context)
        else:
            messages.error(request, "Invalid Form Submitted ")
            return render(request, "hod_template/edit_session_template.html", context)

    else:
        return render(request, "hod_template/edit_session_template.html", context)


@csrf_exempt
def check_email_availability(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Only POST requests are allowed'}, status=405)
    
    email = request.POST.get("email")
    if not email:
        return JsonResponse({'error': 'Email is required'}, status=400)
    
    try:
        # Check if email exists in CustomUser model
        exists = CustomUser.objects.filter(email=email).exists()
        if exists:
            return JsonResponse({
                'is_taken': True,
                'error_message': 'This email is already registered.'
            })
        return JsonResponse({
            'is_taken': False,
            'success_message': 'Email is available.'
        })
    except Exception as e:
        return JsonResponse({
            'error': str(e)
        }, status=500)


@csrf_exempt
def student_feedback_message(request):
    if request.method != 'POST':
        feedbacks = FeedbackStudent.objects.all()
        context = {
            'feedbacks': feedbacks,
            'page_title': 'Student Feedback Messages'
        }
        return render(request, 'hod_template/student_feedback_template.html', context)
    else:
        feedback_id = request.POST.get('id')
        try:
            feedback = get_object_or_404(FeedbackStudent, id=feedback_id)
            reply = request.POST.get('reply')
            feedback.reply = reply
            feedback.save()
            return HttpResponse(True)
        except Exception as e:
            return HttpResponse(False)


@csrf_exempt
def staff_feedback_message(request):
    if request.method != 'POST':
        feedbacks = FeedbackStaff.objects.all()
        context = {
            'feedbacks': feedbacks,
            'page_title': 'Staff Feedback Messages'
        }
        return render(request, 'hod_template/staff_feedback_template.html', context)
    else:
        feedback_id = request.POST.get('id')
        try:
            feedback = get_object_or_404(FeedbackStaff, id=feedback_id)
            reply = request.POST.get('reply')
            feedback.reply = reply
            feedback.save()
            return HttpResponse(True)
        except Exception as e:
            return HttpResponse(False)


@csrf_exempt
def view_staff_leave(request):
    if request.method != 'POST':
        allLeave = LeaveReportStaff.objects.all()
        context = {
            'allLeave': allLeave,
            'page_title': 'Leave Applications From Staff'
        }
        return render(request, "hod_template/staff_leave_view.html", context)
    else:
        id = request.POST.get('id')
        status = request.POST.get('status')
        if (status == '1'):
            status = 1
        else:
            status = -1
        try:
            leave = get_object_or_404(LeaveReportStaff, id=id)
            leave.status = status
            leave.save()
            return HttpResponse(True)
        except Exception as e:
            return False


@csrf_exempt
def view_student_leave(request):
    if request.method != 'POST':
        allLeave = LeaveReportStudent.objects.all()
        context = {
            'allLeave': allLeave,
            'page_title': 'Leave Applications From Students'
        }
        return render(request, "hod_template/student_leave_view.html", context)
    else:
        id = request.POST.get('id')
        status = request.POST.get('status')
        if (status == '1'):
            status = 1
        else:
            status = -1
        try:
            leave = get_object_or_404(LeaveReportStudent, id=id)
            leave.status = status
            leave.save()
            return HttpResponse(True)
        except Exception as e:
            return False


def admin_view_attendance(request):
    subjects = Subject.objects.all()
    sessions = Session.objects.all()
    context = {
        'subjects': subjects,
        'sessions': sessions,
        'page_title': 'View Attendance'
    }

    return render(request, "hod_template/admin_view_attendance.html", context)


@csrf_exempt
def get_admin_attendance(request):
    subject_id = request.POST.get('subject')
    session_id = request.POST.get('session')
    attendance_date_id = request.POST.get('attendance_date_id')
    try:
        subject = get_object_or_404(Subject, id=subject_id)
        session = get_object_or_404(Session, id=session_id)
        attendance = get_object_or_404(
            Attendance, id=attendance_date_id, session=session)
        attendance_reports = AttendanceReport.objects.filter(
            attendance=attendance)
        json_data = []
        for report in attendance_reports:
            data = {
                "status":  str(report.status),
                "name": str(report.student)
            }
            json_data.append(data)
        return JsonResponse(json.dumps(json_data), safe=False)
    except Exception as e:
        return None


@csrf_exempt
def save_attendance_data(request):
    if request.method == 'POST':
        subject_id = request.POST.get('subject_id')
        session_id = request.POST.get('session_id')
        attendance_date = request.POST.get('attendance_date')
        student_ids = request.POST.getlist('student_ids[]')

        logging.info(f"Subject ID: {subject_id}")
        logging.info(f"Session ID: {session_id}")
        logging.info(f"Attendance Date: {attendance_date}")
        logging.info(f"Student IDs: {student_ids}")

        try:
            subject = get_object_or_404(Subject, id=subject_id)
            session = get_object_or_404(Session, id=session_id)

            # Convert attendance_date to a date object
            attendance_date_obj = datetime.strptime(attendance_date, '%Y-%m-%d').date()

            # Validate that the attendance date's year is within the session's academic years
            attendance_year = attendance_date_obj.year
            session_start_year = session.start_year.year if isinstance(session.start_year, datetime) else session.start_year
            session_end_year = session.end_year.year if isinstance(session.end_year, datetime) else session.end_year
            
            if not (session_start_year <= attendance_year <= session_end_year):
                return JsonResponse({
                    "status": "error",
                    "message": f"Attendance year {attendance_year} must be between academic session years "
                              f"{session_start_year} and {session_end_year}"
                })

            # Validate that the attendance date is precisely within the session date range
            if attendance_date_obj < session.start_year.date() or attendance_date_obj > session.end_year.date():
                return JsonResponse({
                    "status": "error",
                    "message": f"Attendance date ({attendance_date_obj.strftime('%Y-%m-%d')}) is outside the session period ({session.start_year.date().strftime('%Y-%m-%d')} - {session.end_year.date().strftime('%Y-%m-%d')})"
                })

            # Check if attendance already exists for the given date and student
            for student_id in student_ids:
                student = get_object_or_404(Student, id=student_id)
                if AttendanceReport.objects.filter(student=student, attendance__date=attendance_date_obj).exists():
                    return JsonResponse({"status": "error", "message": f"Attendance already recorded for student {student.admin.first_name} {student.admin.last_name} on {attendance_date}"})

            attendance = Attendance.objects.create(subject=subject, session=session, date=attendance_date_obj)

            # Check if attendance already exists for the given date and student
            for student_id in student_ids:
                student = get_object_or_404(Student, id=student_id)
                if AttendanceReport.objects.filter(student=student, attendance__date=attendance_date_obj).exists():
                    return JsonResponse({"status": "error", "message": f"Attendance already recorded for student {student.admin.first_name} {student.admin.last_name} on {attendance_date}"})
                try:
                    student = get_object_or_404(Student, id=student_id)
                    status = request.POST.get(f'status_{student_id}', 'False') == 'True'
                    AttendanceReport.objects.create(student=student, attendance=attendance, status=status)
                except Exception as e:
                    logging.error(f"Error processing student ID {student_id}: {str(e)}")
                    messages.error(request, f"Error processing student ID {student_id}: {str(e)}")
                    return JsonResponse({"status": "error", "message": f"Error processing student ID {student_id}: {str(e)}"})

            messages.success(request, "Attendance saved successfully!")
            return JsonResponse({"status": "success"})
        except Exception as e:
            logging.error(f"Error in saving attendance: {str(e)}")
            messages.error(request, f"Error in saving attendance: {str(e)}")
            return JsonResponse({"status": "error", "message": str(e)})

    return JsonResponse({"status": "error", "message": "Invalid request method"})


def admin_view_profile(request):
    admin = get_object_or_404(Admin, admin=request.user)
    form = AdminForm(request.POST or None, request.FILES or None,
                     instance=admin)
    context = {'form': form,
               'page_title': 'View/Edit Profile'
               }
    if request.method == 'POST':
        try:
            if form.is_valid():
                first_name = form.cleaned_data.get('first_name')
                last_name = form.cleaned_data.get('last_name')
                password = form.cleaned_data.get('password') or None
                passport = request.FILES.get('profile_pic') or None
                custom_user = admin.admin
                if password != None:
                    custom_user.set_password(password)
                if passport != None:
                    fs = FileSystemStorage()
                    filename = fs.save(passport.name, passport)
                    passport_url = fs.url(filename)
                    custom_user.profile_pic = passport_url
                custom_user.first_name = first_name
                custom_user.last_name = last_name
                custom_user.save()
                messages.success(request, "Profile Updated!")
                return redirect(reverse('admin_view_profile'))
            else:
                messages.error(request, "Invalid Data Provided")
        except Exception as e:
            messages.error(
                request, "Error Occured While Updating Profile " + str(e))
    return render(request, "hod_template/admin_view_profile.html", context)


def admin_notify_staff(request):
    staff = CustomUser.objects.filter(user_type=2)
    context = {
        'page_title': "Send Notifications To Staff",
        'allStaff': staff
    }
    return render(request, "hod_template/staff_notification.html", context)


def admin_notify_student(request):
    student = CustomUser.objects.filter(user_type=3)
    context = {
        'page_title': "Send Notifications To Students",
        'students': student
    }
    return render(request, "hod_template/student_notification.html", context)


@csrf_exempt
def send_student_notification(request):
    id = request.POST.get('id')
    message = request.POST.get('message')
    student = get_object_or_404(Student, admin_id=id)
    try:
        url = "https://fcm.googleapis.com/fcm/send"
        body = {
            'notification': {
                'title': "Student Management System",
                'body': message,
                'click_action': reverse('student_view_notification'),
                'icon': static('dist/img/AdminLTELogo.png')
            },
            'to': student.admin.fcm_token
        }
        headers = {'Authorization':
                   'key=AAAA3Bm8j_M:APA91bElZlOLetwV696SoEtgzpJr2qbxBfxVBfDWFiopBWzfCfzQp2nRyC7_A2mlukZEHV4g1AmyC6P_HonvSkY2YyliKt5tT3fe_1lrKod2Daigzhb2xnYQMxUWjCAIQcUexAMPZePB',
                   'Content-Type': 'application/json'}
        data = requests.post(url, data=json.dumps(body), headers=headers)
        notification = NotificationStudent(student=student, message=message)
        notification.save()
        return HttpResponse("True")
    except Exception as e:
        return HttpResponse("False")


@csrf_exempt
def send_staff_notification(request):
    id = request.POST.get('id')
    message = request.POST.get('message')
    staff = get_object_or_404(Staff, admin_id=id)
    try:
        url = "https://fcm.googleapis.com/fcm/send"
        body = {
            'notification': {
                'title': "Student Management System",
                'body': message,
                'click_action': reverse('staff_view_notification'),
                'icon': static('dist/img/AdminLTELogo.png')
            },
            'to': staff.admin.fcm_token
        }
        headers = {'Authorization':
                   'key=AAAA3Bm8j_M:APA91bElZlOLetwV696SoEtgzpJr2qbxBfxVBfDWFiopBWzfCfzQp2nRyC7_A2mlukZEHV4g1AmyC6P_HonvSkY2YyliKt5tT3fe_1lrKod2Daigzhb2xnYQMxUWjCAIQcUexAMPZePB',
                   'Content-Type': 'application/json'}
        data = requests.post(url, data=json.dumps(body), headers=headers)
        notification = NotificationStaff(staff=staff, message=message)
        notification.save()
        return HttpResponse("True")
    except Exception as e:
        return HttpResponse("False")


def delete_staff(request, staff_id):
    """
    Delete a staff member and all their related records safely using a transaction.
    Handles updating subjects and cascade deletion of feedback, leave reports, etc.
    
    Args:
        request: HTTP request object
        staff_id: ID of the staff member to delete
        
    Returns:
        Redirects to manage_staff view with success/error message
    """
    staff = get_object_or_404(CustomUser, staff__id=staff_id)
    
    try:
        with transaction.atomic():
            # Store staff name before deletion for message
            staff_name = f"{staff.first_name} {staff.last_name}"
            
            # Get all subjects taught by this staff
            staff_subjects = Subject.objects.filter(staff=staff.staff)
            
            # Delete attendance records for subjects taught by this staff
            attendance_records = Attendance.objects.filter(subject__in=staff_subjects)
            AttendanceReport.objects.filter(attendance__in=attendance_records).delete()
            attendance_records.delete()
            
            # Delete notifications, feedback, and leave reports
            NotificationStaff.objects.filter(staff=staff.staff).delete()
            FeedbackStaff.objects.filter(staff=staff.staff).delete()
            LeaveReportStaff.objects.filter(staff=staff.staff).delete()
            
            # Delete subjects taught by staff
            staff_subjects.delete()
            
            # Delete the staff user which will cascade delete Staff model
            staff.delete()
            
            messages.success(request, f"Staff member '{staff_name}' and all related records deleted successfully!")
            
    except Exception as e:
        messages.error(request, f"Error deleting staff member: {str(e)}")
        
    return redirect(reverse('manage_staff'))

def delete_student(request, student_id):
    """
    Delete a student and all their related records after checking dependencies.
    Handles cascade deletion of attendance records and leave reports.
    """
    student = get_object_or_404(CustomUser, student__id=student_id)
    
    try:
        # Begin transaction
        with transaction.atomic():
            # Delete related records first
            AttendanceReport.objects.filter(student=student.student).delete()
            LeaveReportStudent.objects.filter(student=student.student).delete()
            FeedbackStudent.objects.filter(student=student.student).delete()
            NotificationStudent.objects.filter(student=student.student).delete()
            
            # Delete the student and user
            student_name = f"{student.first_name} {student.last_name}"
            student.delete()
            
            messages.success(request, f"Student '{student_name}' and all related records deleted successfully!")
            
    except Exception as e:
        messages.error(request, f"Error deleting student: {str(e)}")
        
    return redirect(reverse('manage_student'))

def delete_course(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    try:
        course.delete()
        messages.success(request, "Course deleted successfully!")
    except Exception:
        messages.error(
            request, "Sorry, some students are assigned to this course already. Kindly change the affected student course and try again")
    return redirect(reverse('manage_course'))


def delete_subject(request, subject_id):
    """
    Delete a subject safely after checking for dependencies.
    Prevents IntegrityError from foreign key constraints.
    """
    subject = get_object_or_404(Subject, id=subject_id)
    
    # Check for related StudentSubject records
    student_subject_count = StudentSubject.objects.filter(subject=subject).count()
    
    # Check for related Attendance records
    attendance_count = Attendance.objects.filter(subject=subject).count()
    
    # If there are related records, don't allow deletion
    if student_subject_count > 0 or attendance_count > 0:
        # Create a detailed message about the constraints
        constraint_details = []
        if student_subject_count > 0:
            constraint_details.append(f"{student_subject_count} students are enrolled in this subject")
        if attendance_count > 0:
            constraint_details.append(f"{attendance_count} attendance records exist for this subject")
        
        message = "Cannot delete subject due to existing records: " + ", ".join(constraint_details)
        message += ". Please reassign or delete these records first."
        
        messages.error(request, message)
    else:
        try:
            # Safe to delete
            subject_name = subject.name
            subject.delete()
            messages.success(request, f"Subject '{subject_name}' deleted successfully!")
        except Exception as e:
            messages.error(request, f"Error deleting subject: {str(e)}")
    
    return redirect(reverse('manage_subject'))


def delete_session(request, session_id):
    session = get_object_or_404(Session, id=session_id)
    if Student.objects.filter(session=session).exists():
        messages.error(request, "There are students assigned to this session. Please move them to another session.")
    else:
        session.delete()
        messages.success(request, "Session deleted successfully!")
    return redirect(reverse('manage_session'))

# New function to get subjects for a course via AJAX
def get_subjects_for_course(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Invalid request method'}, status=400)
    
    course_id = request.POST.get('course_id')
    if not course_id:
        return JsonResponse({'error': 'Course ID is required'}, status=400)
    
    try:
        # Get subjects for the course
        subjects = Subject.objects.filter(course_id=course_id).values('id', 'name')
        return JsonResponse({'subjects': list(subjects)})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def generate_qr_codes(request):
    """Generate QR codes for students who don't have one yet"""
    
    if request.method == 'POST':
        try:
            # Get all students without QR codes
            students_without_qr = CustomUser.objects.filter(user_type=3, qr_code__isnull=True) | CustomUser.objects.filter(user_type=3, qr_code='')
            
            success_count = 0
            fail_count = 0
            
            for user in students_without_qr:
                try:
                    if qrcode is not None and hasattr(user, 'student'):
                        # If student doesn't have a student code, generate one
                        if not user.student_code:
                            user.student_code = generate_unique_student_code()
                            user.save()
                        
                        # Create QR code with student data - only include student_code
                        student_data = {
                            'student_code': user.student_code
                        }
                        
                        # Create QR code
                        qr = qrcode.QRCode(
                            version=1,
                            error_correction=qrcode.constants.ERROR_CORRECT_L,
                            box_size=10,
                            border=4,
                        )
                        qr.add_data(json.dumps(student_data))
                        qr.make(fit=True)
                        qr_image = qr.make_image(fill_color="black", back_color="white")
                        
                        # Save QR code
                        buffer = BytesIO()
                        qr_image.save(buffer, format='PNG')
                        buffer.seek(0)
                        
                        # Create a unique filename for the QR code
                        filename = f'qr_code_{user.student_code}.png'
                        file_path = os.path.join('qr_codes', filename)
                        
                        # Use FileSystemStorage to save the file
                        fs = FileSystemStorage()
                        saved_path = fs.save(file_path, ContentFile(buffer.getvalue()))
                        
                        # Update user's qr_code field with the saved file path
                        user.qr_code = saved_path
                        user.save()
                        
                        success_count += 1
                    else:
                        fail_count += 1
                except Exception as e:
                    fail_count += 1
                    continue
            
            if success_count > 0:
                messages.success(request, f"Successfully generated {success_count} QR codes.")
            if fail_count > 0:
                messages.warning(request, f"Failed to generate {fail_count} QR codes.")
            
            return redirect(reverse('manage_student'))
            
        except Exception as e:
            messages.error(request, f"Error generating QR codes: {str(e)}")
            return redirect(reverse('manage_student'))
    
    # If not POST, just redirect back to student management
    return redirect(reverse('manage_student'))

def delete_multiple_students(request):
    """
    Delete multiple students based on submitted form data
    """
    if request.method == 'POST':
        selected_ids = request.POST.getlist('selected_students')
        
        if not selected_ids:
            messages.error(request, "No students selected for deletion")
            return redirect(reverse('manage_student'))
        
        try:
            # Begin transaction for safe bulk deletion
            with transaction.atomic():
                deleted_count = 0
                for student_id in selected_ids:
                    try:
                        student = get_object_or_404(CustomUser, student__id=student_id)
                        
                        # Delete related records first
                        AttendanceReport.objects.filter(student=student.student).delete()
                        LeaveReportStudent.objects.filter(student=student.student).delete()
                        FeedbackStudent.objects.filter(student=student.student).delete()
                        NotificationStudent.objects.filter(student=student.student).delete()
                        StudentSubject.objects.filter(student=student.student).delete()
                        
                        # Delete the student and user
                        student.delete()
                        deleted_count += 1
                    except Exception as e:
                        continue
                
                messages.success(request, f"Successfully deleted {deleted_count} student(s)")
        except Exception as e:
            messages.error(request, f"Error during bulk deletion: {str(e)}")
    
    return redirect(reverse('manage_student'))
